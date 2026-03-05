"""
Módulo para gerenciar template e histórico de relatório de horas
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import shutil
from datetime import datetime

TEMPLATE_DIR = Path("templates")
TEMPLATE_DIR.mkdir(exist_ok=True)

TEMPLATE_BASE_PATH = TEMPLATE_DIR / "template_base.xlsx"
LAST_EXPORT_PATH = TEMPLATE_DIR / "last_export.xlsx"

def setup_template(template_file_path):
    """Copia e salva o template base"""
    if Path(template_file_path).exists():
        shutil.copy(template_file_path, TEMPLATE_BASE_PATH)
        return True
    return False

def get_template_base():
    """Retorna o caminho do template base ou None"""
    if TEMPLATE_BASE_PATH.exists():
        return TEMPLATE_BASE_PATH
    # Se não existir, procura na raiz
    template_files = list(Path(".").glob("*Template*.xlsx"))
    if template_files:
        return setup_template(template_files[0])
    return None

def add_new_data_to_template(new_data_df, template_path=None):
    """
    Adiciona novos dados à aba 'Time Reporting Export' do template
    
    Parâmetros:
    - new_data_df: DataFrame com dados do ClickUp (colunas: data, horas, projeto, Task Name, etc)
    - template_path: Caminho do template (ou None para usar o salvo)
    """
    if template_path is None:
        template_path = get_template_base()
    
    if not template_path:
        raise ValueError("❌ Nenhum template encontrado!")
    
    # Carregar o template
    wb = openpyxl.load_workbook(template_path)
    ws_time = wb["Time Reporting Export"]
    
    # Carregar dados existentes
    try:
        df_existing = pd.read_excel(template_path, sheet_name="Time Reporting Export")
    except:
        df_existing = pd.DataFrame()
    
    # Preparar novos dados para adicionar
    # Precisa mapear os dados do ClickUp para o formato do template
    new_data_processed = prepare_data_for_template(new_data_df)
    
    # Combinar dados
    if len(df_existing) > 0:
        df_combined = pd.concat([df_existing, new_data_processed], ignore_index=True)
    else:
        df_combined = new_data_processed
    
    # Remover duplicatas baseado em "Time Entry ID" se existir
    if "Time Entry ID" in df_combined.columns:
        df_combined = df_combined.drop_duplicates(subset=["Time Entry ID"], keep="last")
    
    return df_combined, wb

def prepare_data_for_template(df_clickup):
    """
    Converte DataFrame do ClickUp para o formato esperado pelo template
    Mapeia as colunas necessárias
    """
    # Colunas esperadas pelo template
    template_columns = [
        'User ID', 'Username', 'Time Entry ID', 'Description', 'Billable',
        'Time Labels', 'Start', 'Start Text', 'Stop', 'Stop Text',
        'Time Tracked', 'Time Tracked Text', 'Space ID', 'Space Name',
        'Folder ID', 'Folder Name', 'List ID', 'List Name', 'Task ID',
        'Task Name', 'Task Status', 'Due Date', 'Due Date Text',
        'Start Date', 'Start Date Text', 'Task Time Estimated',
        'Task Time Estimated Text', 'Task Time Spent', 'Task Time Spent Text',
        'User Total Time Estimated', 'User Total Time Estimated Text',
        'User Total Time Tracked', 'User Total Time Tracked Text',
        'Tags', 'Checklists', 'User Period Time Spent',
        'User Period Time Spent Text', 'Date Created', 'Date Created Text',
        'Custom Task ID', 'Parent Task ID', 'Remaining Hours',
        'Hour Utilization (%)', 'Hour Utilization Alert', 'Billable.1',
        'Client', 'User', 'Parent Task Name', 'Task Name (Report)',
        'Duration (Decimal)', 'Duration', 'Task Date', 'Year_Month',
        'Reference', 'Week of the Year', 'Month', 'Month Abbreviation',
        'Week of the Month'
    ]
    
    # Se o DataFrame já tem as colunas do template, apenas retorna
    if all(col in df_clickup.columns for col in ['Time Entry ID', 'Task Name', 'Start Text', 'Time Tracked Text']):
        # Completar colunas faltantes com NaN
        for col in template_columns:
            if col not in df_clickup.columns:
                df_clickup[col] = None
        return df_clickup[template_columns]
    
    # Caso contrário, precisa fazer o mapeamento
    df_processed = pd.DataFrame()
    
    # Copiar todas as colunas disponíveis
    for col in df_clickup.columns:
        if col in template_columns:
            df_processed[col] = df_clickup[col]
    
    # Adicionar colunas faltantes
    for col in template_columns:
        if col not in df_processed.columns:
            df_processed[col] = None
    
    return df_processed[template_columns]

def export_with_template(df_combined, wb, output_filename):
    """
    Exporta o arquivo mantendo a estrutura do template
    Atualiza apenas a aba 'Time Reporting Export' com os dados novos
    """
    # Limpar a aba Time Reporting Export
    ws_time = wb["Time Reporting Export"]
    
    # Manter apenas a linha de header e limpar o resto
    if ws_time.max_row > 1:
        ws_time.delete_rows(2, ws_time.max_row)
    
    # Escrever novos dados
    for r_idx, row in enumerate(df_combined.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws_time.cell(row=r_idx, column=c_idx, value=value)
    
    # Salvar arquivo
    wb.save(output_filename)
    return output_filename

def merge_with_history(new_data, use_last_export=True):
    """
    Combine novos dados com o histórico anterior
    """
    try:
        # Tentar carregar último export
        if use_last_export and LAST_EXPORT_PATH.exists():
            df_last = pd.read_excel(LAST_EXPORT_PATH, sheet_name="Time Reporting Export")
        else:
            df_last = pd.DataFrame()
    except:
        df_last = pd.DataFrame()
    
    # Combinar dados
    if len(df_last) > 0:
        df_combined = pd.concat([df_last, new_data], ignore_index=True)
        # Remover duplicatas
        if "Time Entry ID" in df_combined.columns:
            df_combined = df_combined.drop_duplicates(subset=["Time Entry ID"], keep="last")
    else:
        df_combined = new_data
    
    return df_combined
